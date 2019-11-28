# -*- coding: utf-8 -*-
import os
import re
import datetime
import traceback
import subprocess
from openpyxl import load_workbook
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import View, ListView, TemplateView
from django.contrib.auth.mixins import PermissionRequiredMixin
from django.db.models import Q, F
from django.http import HttpResponse
from wsgiref.util import FileWrapper

from common.mixins import JSONResponseMixin
from common.utils.base import BASE_DIR

from cmdb.models.base import IDC
from cmdb.models.accessory import accessory_item, CPU, Memory, Disk, Caddy, NetworkAdapter, NetworkCable, \
    OpticalTransceiver, JumpWire, Accessory, InventoryRecord


def get_accessory_version(accessory, search):
    result = list()
    if accessory == 'cpu':
        if search:
            obj_list = CPU.objects.filter(Q(version__contains=search) |
                                          Q(speed=search) | Q(process=search)).distinct()
        else:
            obj_list = CPU.objects.get_queryset()
        for obj in obj_list:
            acc_list = Accessory.objects.filter(mode=accessory, mode_id=obj.id)
            cpu_used = acc_list.filter(is_active=False).count()
            cpu_available = acc_list.filter(is_active=True).count()
            result.append({
                'id': obj.id,
                'accessory': accessory,
                'accessory_name': accessory_item[accessory],
                'version': '%s|%dMHz|%d核' % (obj.version, obj.speed, obj.process),
                'total': cpu_used + cpu_available,
                'residual': cpu_available,
                'created_date': obj.created_date
            })
    elif accessory == 'memory':
        if search:
            obj_list = Memory.objects.filter(Q(ram_type=search) |
                                             Q(ram_size=search) | Q(speed=search)).distinct()
        else:
            obj_list = Memory.objects.get_queryset()
        for obj in obj_list:
            acc_list = Accessory.objects.filter(mode=accessory, mode_id=obj.id)
            cpu_used = acc_list.filter(is_active=False).count()
            cpu_available = acc_list.filter(is_active=True).count()
            result.append({
                'id': obj.id,
                'accessory': accessory,
                'accessory_name': accessory_item[accessory],
                'version': '%s|%dG|%dMT/s' % (obj.get_ram_type_display(), obj.ram_size, obj.speed),
                'total': cpu_used + cpu_available,
                'residual': cpu_available,
                'created_date': obj.created_date
            })
    elif accessory == 'disk':
        if search:
            obj_list = Disk.objects.filter(Q(device_type=search) | Q(rpm=search) |
                                           Q(capacity=search) | Q(dimensions=search)).distinct()
        else:
            obj_list = Disk.objects.get_queryset()
        for obj in obj_list:
            acc_list = Accessory.objects.filter(mode=accessory, mode_id=obj.id)
            cpu_used = acc_list.filter(is_active=False).count()
            cpu_available = acc_list.filter(is_active=True).count()
            result.append({
                'id': obj.id,
                'accessory': accessory,
                'accessory_name': accessory_item[accessory],
                'version': '%s|%dG|%d转/s|%s' % (
                    obj.get_device_type_display(), obj.capacity, obj.rpm, obj.get_dimensions_display()),
                'total': cpu_used + cpu_available,
                'residual': cpu_available,
                'created_date': obj.created_date
            })
    elif accessory == 'caddy':
        if search:
            obj_list = Caddy.objects.filter(dimensions=search)
        else:
            obj_list = Caddy.objects.get_queryset()
        for obj in obj_list:
            acc_list = Accessory.objects.filter(mode=accessory, mode_id=obj.id)
            cpu_used = acc_list.filter(is_active=False).count()
            cpu_available = acc_list.filter(is_active=True).count()
            result.append({
                'id': obj.id,
                'accessory': accessory,
                'accessory_name': accessory_item[accessory],
                'version': obj.get_dimensions_display(),
                'total': cpu_used + cpu_available,
                'residual': cpu_available,
                'created_date': obj.created_date
            })
    elif accessory == 'network_adapter':
        if search:
            obj_list = NetworkAdapter.objects.filter(speed=search)
        else:
            obj_list = NetworkAdapter.objects.get_queryset()
        for obj in obj_list:
            acc_list = Accessory.objects.filter(mode=accessory, mode_id=obj.id)
            cpu_used = acc_list.filter(is_active=False).count()
            cpu_available = acc_list.filter(is_active=True).count()
            result.append({
                'id': obj.id,
                'accessory': accessory,
                'accessory_name': accessory_item[accessory],
                'version': obj.get_speed_display(),
                'total': cpu_used + cpu_available,
                'residual': cpu_available,
                'created_date': obj.created_date
            })
    elif accessory == 'network_cable':
        if search:
            obj_list = NetworkCable.objects.filter(Q(cat=search) | Q(length=search)).distinct()
        else:
            obj_list = NetworkCable.objects.get_queryset()
        for obj in obj_list:
            acc_list = Accessory.objects.filter(mode=accessory, mode_id=obj.id)
            cpu_used = acc_list.filter(is_active=False).count()
            cpu_available = acc_list.filter(is_active=True).count()
            result.append({
                'id': obj.id,
                'accessory': accessory,
                'accessory_name': accessory_item[accessory],
                'version': '%s|%d米' % (obj.get_cat_display(), obj.length),
                'total': cpu_used + cpu_available,
                'residual': cpu_available,
                'created_date': obj.created_date
            })
    elif accessory == 'transceiver':
        if search:
            obj_list = OpticalTransceiver.objects.filter(
                Q(information__contains=search) | Q(mode=search) | Q(reach=search) | Q(rate=search)).distinct()
        else:
            obj_list = OpticalTransceiver.objects.get_queryset()
        for obj in obj_list:
            acc_list = Accessory.objects.filter(mode=accessory, mode_id=obj.id)
            cpu_used = acc_list.filter(is_active=False).count()
            cpu_available = acc_list.filter(is_active=True).count()
            result.append({
                'id': obj.id,
                'accessory': accessory,
                'accessory_name': accessory_item[accessory],
                'version': '%s|%s|%dKM|%s' % (
                    obj.information, obj.get_mode_display(), obj.reach, obj.get_rate_display()),
                'total': cpu_used + cpu_available,
                'residual': cpu_available,
                'created_date': obj.created_date
            })
    elif accessory == 'jump_wire':
        if search:
            obj_list = JumpWire.objects.filter(
                Q(information__contains=search) | Q(mode=search) | Q(interface=search) | Q(length=search)).distinct()
        else:
            obj_list = JumpWire.objects.get_queryset()
        for obj in obj_list:
            acc_list = Accessory.objects.filter(mode=accessory, mode_id=obj.id)
            cpu_used = acc_list.filter(is_active=False).count()
            cpu_available = acc_list.filter(is_active=True).count()
            result.append({
                'id': obj.id,
                'accessory': accessory,
                'accessory_name': accessory_item[accessory],
                'version': '%s|%s|%s|%s米' % (
                    obj.information, obj.get_mode_display(), obj.get_interface_display(), obj.length),
                'total': cpu_used + cpu_available,
                'residual': cpu_available,
                'created_date': obj.created_date
            })
    else:
        pass
    return result


class AccessoryView(PermissionRequiredMixin, TemplateView):
    permission_required = 'auth.perm_cmdb_accessory_view'
    template_name = "cmdb/accessory.html"

    def get_context_data(self, **kwargs):
        context = {
            'path1': 'CMDB',
            'path2': '配件管理',
            'acc_item': accessory_item,
            'idc_list': [{"id": i.id, "name": i.idc_name} for i in IDC.objects.get_queryset()]
        }
        kwargs.update(context)
        return super().get_context_data(**kwargs)


class AccessoryListView(PermissionRequiredMixin, JSONResponseMixin, TemplateView):
    permission_required = 'auth.perm_cmdb_accessory_view'

    def get(self, request, **kwargs):
        accessory = request.GET.get("accessory", '')
        search = request.GET.get("search", '')
        result = list()
        if accessory:
            result.extend(get_accessory_version(accessory, search))
        else:
            result.extend(get_accessory_version('cpu', search))
            result.extend(get_accessory_version('memory', search))
            result.extend(get_accessory_version('disk', search))
            result.extend(get_accessory_version('caddy', search))
            result.extend(get_accessory_version('network_adapter', search))
            result.extend(get_accessory_version('network_cable', search))
            result.extend(get_accessory_version('transceiver', search))
            result.extend(get_accessory_version('jump_wire', search))
        return self.render_json_response(result)

    def post(self, request):
        # 配件采购入库
        if not request.user.has_perm('auth.perm_cmdb_accessory_edit'):
            return self.render_json_response({'code': 1, 'errmsg': '权限不足，无法新增！'})
        try:
            acc_list = list()
            idc_id = request.POST.get('idc_id', 0)
            idc = IDC.objects.get(id=idc_id)
            mode = request.POST.get('mode')
            mode_id = request.POST.get('mode_id')
            count = int(request.POST.get('count', 0))
            for _ in range(0, count):
                acc_list.append(Accessory(
                    storehouse=idc,
                    mode=mode,
                    mode_id=mode_id,
                    manufacturer=request.POST.get('manufacturer'),
                    sn=request.POST.get('sn'),
                    vendor=request.POST.get('vendor'),
                    trade_date=request.POST.get('trade_date'),
                    expired_date=request.POST.get('expired_date'),
                    comment=request.POST.get('comment')
                ))
            Accessory.objects.bulk_create(acc_list)
            if count > 0:
                if mode == 'cpu':
                    content = '%s %d 块！' % (CPU.objects.get(id=mode_id).version, count)
                elif mode == 'memory':
                    obj = Memory.objects.get(id=mode_id)
                    version = '%s|%dG|%dMT/s' % (obj.get_ram_type_display(), obj.ram_size, obj.speed)
                    content = '%s %d 块！' % (version, count)
                elif mode == 'disk':
                    obj = Disk.objects.get(id=mode_id)
                    version = '%s|%dG|%d转/s|%s' % (obj.get_device_type_display(), obj.capacity,
                                                   obj.rpm, obj.get_dimensions_display())
                    content = '%s %d 块！' % (version, count)
                elif mode == 'caddy':
                    obj = Caddy.objects.get(id=mode_id)
                    content = '%s 硬盘托架 %d 个！' % (obj.get_dimensions_display(), count)
                elif mode == 'network_adapter':
                    obj = NetworkAdapter.objects.get(id=mode_id)
                    content = '%s 网卡 %d 个！' % (obj.get_speed_display(), count)
                elif mode == 'network_cable':
                    obj = NetworkCable.objects.get(id=mode_id)
                    content = '%s %d米 网线 %s 根！' % (obj.get_cat_display(), obj.length, count)
                elif mode == 'transceiver':
                    obj = OpticalTransceiver.objects.get(id=mode_id)
                    version = '%s|%s|%dKM|%s' % (obj.information, obj.get_mode_display(),
                                                 obj.reach, obj.get_rate_display())
                    content = '%s 光模块 %d 个！' % (version, count)
                elif mode == 'jump_wire':
                    obj = JumpWire.objects.get(id=mode_id)
                    content = '%s %s %d米 跳线 %s 根！' % (
                        obj.get_mode_display(), obj.get_interface_display(), obj.length, count)

                InventoryRecord.objects.create(accessory=mode, operate='purchase', content=content, user=request.user)
            res = {"code": 0, "result": "添加成功"}
        except IDC.DoesNotExist:
            res = {"code": 1, "errmsg": "请选择仓库！"}
        except Exception as e:
            traceback.print_exc()
            res = {"code": 1, "errmsg": "添加失败：%s" % str(e)}
        return self.render_json_response(res)


@method_decorator(csrf_exempt, name='dispatch')
class AccessoryTemplateView(PermissionRequiredMixin, JSONResponseMixin, View):
    permission_required = 'auth.perm_cmdb_accessory_edit'

    def get(self, request, *args, **kwargs):
        try:
            filename = 'accessory_template.xlsx'
            export_file = os.path.join(BASE_DIR, 'cmdb', 'docs', filename)
            if os.path.isfile(export_file):
                wrapper = FileWrapper(open(export_file, "rb"))
                response = HttpResponse(wrapper, content_type='application/vnd.ms-excel')
                response['Content-Length'] = os.path.getsize(export_file)
                response['Content-Disposition'] = 'attachment; filename="{}"'.format(filename)
                return response
            else:
                return HttpResponse('模板文件不存在！')
        except Exception as e:
            return HttpResponse('模板文件下载失败：' + str(e))

    def post(self, request, *args, **kwargs):
        import_files = request.FILES.getlist('files', None)
        if len(import_files) == 0:
            return self.render_json_response({'code': 1, 'msg': '文件为空，或上传错误！'})
        date_now = datetime.datetime.now().strftime("%Y%m%d%H%M%S%f")
        file_dir = os.path.join(BASE_DIR, 'cmdb', 'docs')
        created, updated, failed = [], [], []
        cols = ['acc', 'id', 'idc_name', 'manufacturer', 'sn', 'vendor', 'trade_date', 'expired_date', 'comment', 'count']
        for im_file in import_files:
            file_name = im_file.name
            file_path = os.path.join(file_dir, request.user.username + '_' + date_now + '_' + file_name)
            with open(file_path, 'wb') as f:
                for chunk in im_file.chunks():
                    f.write(chunk)
            if not os.path.isfile(file_path):
                failed.append('{}：文件为空，或上传错误！'.format(file_name))
                continue
            if not re.match(r'application/vnd\.(ms-excel|ms-office)', subprocess.getoutput('file -b --mime-type ' + file_path)):
                failed.append('{}：请上传excel文件!'.format(im_file))
                continue

            wb = load_workbook(file_path)
            ws = wb.active

            for row in ws.iter_rows(min_row=3, max_col=9):
                row_number = row[0].row
                cols_value = list()
                for cell in row:
                    cols_value.append(cell.value)
                print(row_number, dict(zip(cols, cols_value)))

                row = dict(zip(cols, cols_value))
                if not IDC.objects.filter(idc_name=row['idc_name']).exists():
                    failed.append('{} {}行：IDC名称错误，无法找到！'.format(file_name, str(row_number)))
                    continue
                idc = IDC.objects.get(idc_name=row['idc_name'])
                if row['acc'] not in accessory_item:
                    failed.append('{} {}行：配件类型填写错误！'.format(file_name, str(row_number)))
                    continue
                if not row['id']:
                    failed.append('{} {}行：配件id 未填写！'.format(file_name, str(row_number)))
                    continue
                if row['count'] <= 0:
                    failed.append('{} {}行：配件数量必须大于0！'.format(file_name, str(row_number)))
                    continue
                if row['acc'] == 'cpu':
                    if not CPU.objects.filter(id=row['id']).exists():
                        failed.append('{} {}行：不存在该Id的CPU 型号！'.format(file_name, str(row_number)))
                        continue
                elif row['acc'] == 'memory':
                    if not Memory.objects.filter(id=row['id']).exists():
                        failed.append('{} {}行：不存在该Id的内存 型号！'.format(file_name, str(row_number)))
                        continue
                elif row['acc'] == 'disk':
                    if not Disk.objects.filter(id=row['id']).exists():
                        failed.append('{} {}行：不存在该Id的硬盘 型号！'.format(file_name, str(row_number)))
                        continue
                elif row['acc'] == 'caddy':
                    if not Caddy.objects.filter(id=row['id']).exists():
                        failed.append('{} {}行：不存在该Id的硬盘托架 型号！'.format(file_name, str(row_number)))
                        continue
                elif row['acc'] == 'network_adapter':
                    if not NetworkAdapter.objects.filter(id=row['id']).exists():
                        failed.append('{} {}行：不存在该Id的网卡 型号！'.format(file_name, str(row_number)))
                        continue
                elif row['acc'] == 'network_cable':
                    if not NetworkCable.objects.filter(id=row['id']).exists():
                        failed.append('{} {}行：不存在该Id的网线 型号！'.format(file_name, str(row_number)))
                        continue
                elif row['acc'] == 'transceiver':
                    if not OpticalTransceiver.objects.filter(id=row['id']).exists():
                        failed.append('{} {}行：不存在该Id的光模块 型号！'.format(file_name, str(row_number)))
                        continue
                elif row['acc'] == 'jump_wire':
                    if not JumpWire.objects.filter(id=row['id']).exists():
                        failed.append('{} {}行：不存在该Id的跳线 型号！'.format(file_name, str(row_number)))
                        continue
                for _ in range(row['count']):
                    Accessory.objects.create(storehouse=idc, mode=row['acc'], mode_id=row['id'],
                                             manufacturer=row['manufacturer'], sn=row['sn'],
                                             vendor=row['vendor'], trade_date=row['trade_date'],
                                             expired_date=row['expired_date'], comment=row['comment'])
                created.append('{} {}行：入库成功！'.format(file_name, str(row_number)))
        data = {
            'code': 0,
            'created': created,
            'created_info': 'Created {}'.format(len(created)),
            'updated': updated,
            'updated_info': 'Updated {}'.format(len(updated)),
            'failed': failed,
            'failed_info': 'Failed {}'.format(len(failed)),
            'msg': 'Created: {}. Updated: {}, Error: {}'.format(
                len(created), len(updated), len(failed))
        }
        return self.render_json_response(data)
